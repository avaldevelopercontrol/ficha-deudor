"""
Portfolio Control Center - ETAPA 6 / Avance 4

Lee base-goals.xlsx y genera un SQL autocontenido para cargar la meta mensual
CLARO en aval_analytics.

El script NO se conecta a aval_analytics. Esto permite usarlo desde el Ubuntu
actual aunque ese host siga sin conectividad ODBC hacia 172.23.1.180.

Uso:
    python scripts/analytics/export_claro_goals_snapshot_sql.py \
        --input /ruta/base-goals.xlsx \
        --campaign 2026-08

Salida por defecto:
    /tmp/claro_goal_snapshot_2026-08.sql

Regla source canonical V1:
- CARTERA = CLARO
- base-goals.xlsx es un catálogo vigente de metas por cartera; AÑO y ASIGNACIÓN
  son metadata de la fila fuente y NO determinan la campaña Analytics destino
- debe existir una única fila CLARO con META_PAGOS válida en el snapshot del Excel
- esa META_PAGOS se reutiliza para la campaña indicada en --campaign
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except ModuleNotFoundError:
    print(
        "ERROR: falta la dependencia 'openpyxl'. "
        "Ejecuta: pip install -r scripts/analytics/requirements.txt",
        file=sys.stderr,
    )
    raise SystemExit(2)


SOURCE_CODE = "CLARO_BASE_GOALS"
REQUIRED_HEADERS = {"CARTERA", "ASIGNACION", "ANO", "META_PAGOS"}


@dataclass(frozen=True)
class GoalSnapshot:
    campaign_code: str
    campaign_year: int
    campaign_month: int
    target_recovered_amount: Decimal
    source_rows: int
    source_reference: str
    source_as_of_at: datetime


def normalize_text(value: object | None) -> str:
    if value is None:
        return ""

    text = str(value).strip().upper()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text)


def normalize_header(value: object | None) -> str:
    return normalize_text(value).replace(" ", "_")


def parse_campaign(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"(\d{4})-(\d{2})", value.strip())
    if not match:
        raise ValueError("--campaign debe tener formato YYYY-MM.")

    year = int(match.group(1))
    month = int(match.group(2))

    if month < 1 or month > 12:
        raise ValueError("El mes de --campaign debe estar entre 01 y 12.")

    return year, month


def parse_year(value: object | None) -> int | None:
    if value is None or str(value).strip() == "":
        return None

    try:
        year = int(float(str(value).strip()))
    except ValueError:
        return None

    if year < 2000 or year > 2100:
        return None

    return year


def parse_assignment_month(value: object | None) -> int | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        month = int(value)
        return month if 1 <= month <= 12 else None

    text = normalize_text(value)
    matches = re.findall(r"(?<!\d)(0?[1-9]|1[0-2])(?!\d)", text)

    if not matches:
        return None

    month = int(matches[-1])
    return month if 1 <= month <= 12 else None


def parse_decimal(value: object | None) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace("S/", "").replace(" ", "")

    # Soporta formatos comunes 1,234.56 y 1.234,56.
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text and "." not in text:
        right = text.rsplit(",", 1)[1]
        text = text.replace(",", "") if len(right) == 3 else text.replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def iter_excel_table_rows(workbook) -> tuple[str, list[dict[str, object]]]:
    for worksheet in workbook.worksheets:
        for table_name in worksheet.tables:
            if normalize_text(table_name) != "BASEGOALS":
                continue

            table = worksheet.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            values = list(
                worksheet.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                )
            )

            if not values:
                continue

            headers = [normalize_header(value) for value in values[0]]
            rows = [
                dict(zip(headers, row))
                for row in values[1:]
                if any(value is not None for value in row)
            ]
            return f"{worksheet.title}!{table_name}", rows

    raise LookupError("No se encontró la tabla Excel 'baseGoals'.")


def iter_header_rows(workbook) -> tuple[str, list[dict[str, object]]]:
    for worksheet in workbook.worksheets:
        max_scan = min(20, worksheet.max_row)

        for row_number in range(1, max_scan + 1):
            values = [
                worksheet.cell(row=row_number, column=column).value
                for column in range(1, worksheet.max_column + 1)
            ]
            headers = [normalize_header(value) for value in values]

            if not REQUIRED_HEADERS.issubset(set(headers)):
                continue

            rows: list[dict[str, object]] = []
            for data_row in worksheet.iter_rows(
                min_row=row_number + 1,
                max_row=worksheet.max_row,
                values_only=True,
            ):
                if not any(value is not None for value in data_row):
                    continue
                rows.append(dict(zip(headers, data_row)))

            return f"{worksheet.title}!row{row_number}", rows

    raise LookupError(
        "No se encontró la tabla 'baseGoals' ni una hoja con las columnas "
        "CARTERA, ASIGNACIÓN, AÑO y META_PAGOS."
    )


def read_rows(path: Path) -> tuple[str, list[dict[str, object]]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        try:
            return iter_excel_table_rows(workbook)
        except LookupError:
            return iter_header_rows(workbook)
    finally:
        workbook.close()


def build_snapshot(
    path: Path,
    campaign_code: str,
    source_as_of_at: datetime,
) -> GoalSnapshot:
    campaign_year, campaign_month = parse_campaign(campaign_code)
    source_location, rows = read_rows(path)

    matched: list[tuple[dict[str, object], Decimal]] = []

    for row in rows:
        if normalize_text(row.get("CARTERA")) != "CLARO":
            continue

        amount = parse_decimal(row.get("META_PAGOS"))
        if amount is None:
            continue

        if amount < 0:
            raise ValueError("META_PAGOS no puede ser negativa para CLARO.")

        matched.append((row, amount))

    if not matched:
        raise ValueError(
            "No se encontró una fila CARTERA=CLARO con META_PAGOS válida "
            "en el snapshot actual de base-goals.xlsx."
        )

    if len(matched) != 1:
        raise ValueError(
            "Se encontraron múltiples filas CARTERA=CLARO con META_PAGOS válida. "
            "La regla V1 reutiliza una única meta vigente por cartera; no se "
            "sumarán ni elegirán filas ambiguas automáticamente."
        )

    row, amount = matched[0]
    source_year = parse_year(row.get("ANO"))
    source_month = parse_assignment_month(row.get("ASIGNACION"))
    source_assignment = str(row.get("ASIGNACION") or "").strip()

    source_metadata: list[str] = []
    if source_year is not None:
        source_metadata.append(f"source_year={source_year}")
    if source_assignment:
        source_metadata.append(f"source_assignment={source_assignment}")
    elif source_month is not None:
        source_metadata.append(f"source_assignment_month={source_month:02d}")

    reference = f"{path.name}:{source_location}"
    if source_metadata:
        reference += "; " + "; ".join(source_metadata)

    return GoalSnapshot(
        campaign_code=campaign_code,
        campaign_year=campaign_year,
        campaign_month=campaign_month,
        target_recovered_amount=amount.quantize(Decimal("0.0001")),
        source_rows=1,
        source_reference=reference,
        source_as_of_at=source_as_of_at,
    )


def sql_string(value: str) -> str:
    return "N'" + value.replace("'", "''") + "'"


def sql_datetime(value: datetime) -> str:
    return "'" + value.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "'"


def build_sql(snapshot: GoalSnapshot, crm_client_id: int) -> str:
    source_reference = snapshot.source_reference[:200]

    return f"""/*
Portfolio Control Center - ETAPA 6 / Avance 4
Meta mensual CLARO exportada desde base-goals.xlsx
campaign_code: {snapshot.campaign_code}
source_rows: {snapshot.source_rows}
target_recovered_amount: {snapshot.target_recovered_amount}
source_as_of_at: {snapshot.source_as_of_at}

EJECUTAR EN SSMS CONTRA aval_analytics DESPUÉS DE 011_portfolio_v1_claro_target_support.sql
*/

USE aval_analytics;
GO

SET NOCOUNT ON;
SET XACT_ABORT ON;
GO

BEGIN TRY
    BEGIN TRANSACTION;

    DELETE FROM staging.claro_goal_monthly
    WHERE source_code = '{SOURCE_CODE}'
      AND campaign_code = '{snapshot.campaign_code}';

    INSERT INTO staging.claro_goal_monthly
    (
        source_code,
        campaign_code,
        campaign_year,
        campaign_month,
        target_recovered_amount,
        source_rows,
        source_reference,
        source_as_of_at
    )
    VALUES
    (
        '{SOURCE_CODE}',
        '{snapshot.campaign_code}',
        {snapshot.campaign_year},
        {snapshot.campaign_month},
        {snapshot.target_recovered_amount},
        {snapshot.source_rows},
        {sql_string(source_reference)},
        {sql_datetime(snapshot.source_as_of_at)}
    );

    EXEC etl.usp_load_claro_target_monthly
        @crm_client_id = {crm_client_id},
        @campaign_code = '{snapshot.campaign_code}',
        @source_code = '{SOURCE_CODE}';

    IF NOT EXISTS
    (
        SELECT 1
        FROM analytics.fact_target_monthly AS t
        INNER JOIN analytics.dim_client AS cli
            ON cli.client_key = t.client_key
        INNER JOIN analytics.dim_campaign AS c
            ON c.campaign_key = t.campaign_key
        WHERE cli.crm_client_id = {crm_client_id}
          AND c.campaign_code = '{snapshot.campaign_code}'
          AND t.portfolio_key IS NULL
          AND t.target_recovered_amount = {snapshot.target_recovered_amount}
          AND t.source_as_of_at = {sql_datetime(snapshot.source_as_of_at)}
    )
        THROW 52150, 'La meta cargada no coincide con el snapshot exportado.', 1;

    COMMIT TRANSACTION;

    SELECT
        '{snapshot.campaign_code}' AS campaign_code,
        CAST({snapshot.target_recovered_amount} AS DECIMAL(19,4)) AS target_recovered_amount,
        {snapshot.source_rows} AS source_rows,
        {sql_datetime(snapshot.source_as_of_at)} AS source_as_of_at,
        'CLARO_TARGET_IMPORT_OK' AS assessment;
END TRY
BEGIN CATCH
    IF @@TRANCOUNT > 0
        ROLLBACK TRANSACTION;

    THROW;
END CATCH;
GO
"""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Ruta a base-goals.xlsx")
    parser.add_argument(
        "--campaign",
        required=True,
        help="Campaña Analytics en formato YYYY-MM, por ejemplo 2026-08.",
    )
    parser.add_argument(
        "--crm-client-id",
        type=int,
        default=95,
        help="crm_client_id de CLARO en Analytics. Default: 95.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Archivo SQL de salida. Default: /tmp/claro_goal_snapshot_<campaign>.sql",
    )
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"No existe el archivo: {input_path}")

    campaign_year, campaign_month = parse_campaign(args.campaign)
    campaign_code = f"{campaign_year:04d}-{campaign_month:02d}"
    source_as_of_at = datetime.now()
    source_as_of_at = source_as_of_at.replace(
        microsecond=(source_as_of_at.microsecond // 1000) * 1000
    )

    snapshot = build_snapshot(
        input_path,
        campaign_code,
        source_as_of_at,
    )

    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else Path(f"/tmp/claro_goal_snapshot_{campaign_code}.sql")
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        build_sql(snapshot, args.crm_client_id),
        encoding="utf-8",
    )

    print(f"Fuente: {snapshot.source_reference}")
    print(f"campaign_code: {snapshot.campaign_code}")
    print(f"source_rows: {snapshot.source_rows}")
    print(f"target_recovered_amount: {snapshot.target_recovered_amount}")
    print(f"source_as_of_at: {snapshot.source_as_of_at}")
    print(f"SQL generado: {output}")
    print("")
    print("Ejecuta ese SQL en SSMS contra aval_analytics.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001 - CLI debe terminar con mensaje corto
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
